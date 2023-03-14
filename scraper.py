import os
import sys
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
import time
from selenium.webdriver.common.by import By
import openpyxl
import requests

def main():
    BrowserPath=ResourcePath("./browser/chrome.exe") # ブラウザ
    DriverPath=ResourcePath("./driver/chromedriver.exe") # ウェブドライバ

    # ウェブドライバ設定
    options=Options()
    options.binary_location=BrowserPath
    # options.add_argument("--headless") # 動きを見たい場合はコメントアウトする。
    driver=webdriver.Chrome(DriverPath, options=options)

    # エクセル準備
    path = './ms_agent_list.xlsx'
    wb = openpyxl.load_workbook(path)
    ws = wb["list1"]
    rownumber = 2

    # スクレイピング準備
    url = 'https://www.jmsc.co.jp/entry/list/?issubmit=1&jobgroup%5B1%5D=1&job%5Bjob_type_id%5D%5B98%5D=98&job%5Bjob_type_id%5D%5B99%5D=99&job%5Bjob_type_id%5D%5B100%5D=100&job%5Bjob_type_id%5D%5B119%5D=119&job%5Bincome_year_from%5D=&job%5Bincome_year_to%5D='
    driver.get(url)
    time.sleep(2)
    ttlcnt = 0

    # ページ数
    pagecnt = 0

    # スクレイピング開始ページ
    targetpage = 111
    if targetpage != 1:
        for j in range(1, targetpage):
            nextbutton = driver.find_elements(By.CLASS_NAME, "link_next")
            if len(nextbutton) > 1:
                nextbutton[0].click()

    # 行数
    rownumber = 2

    while True:

        pagecnt += 1
        print(pagecnt)

        # 変数リセット
        titles = []
        wkplaces = []
        positiones = []
        incomes = []

        # リスト数取得
        jobs = driver.find_elements(By.CLASS_NAME, "seminar-list-detail")
        jobscnt = len(jobs)

        # タイトル取得
        for i in range(2, jobscnt + 2):
            wktitles = driver.find_elements(By.XPATH, "//*[@id='entry']/div[1]/div[2]/div[1]/section[2]/div["+str(i)+"]/table/tbody/tr[1]/td/a/h2")
            for wktitle in wktitles:
                if len(wktitle.text) != 0:
                    titles.append(wktitle.text)

        # ポジション取得
        for i in range(2, jobscnt + 2):
            wkpositiones = driver.find_elements(By.XPATH,"//*[@id='entry']/div[1]/div[2]/div[1]/section[2]/div["+str(i)+"]/table/tbody/tr[2]/td")
            for wkposition in wkpositiones:
                if len(wkposition.text) != 0:
                    positiones.append(wkposition.text)

        # 勤務地
        for i in range(2, jobscnt + 2):
            wkwkplaces = driver.find_elements(By.XPATH,"//*[@id='entry']/div[1]/div[2]/div[1]/section[2]/div["+str(i)+"]/table/tbody/tr[4]/td")
            for wkwkplace in wkwkplaces:
                if len(wkwkplace.text) != 0:
                    wkplaces.append(wkwkplace.text)
                else:
                    wkplaces.append('')

        # 年収
        for i in range(2, jobscnt + 2):
            wkincomes = driver.find_elements(By.XPATH,"//*[@id='entry']/div[1]/div[2]/div[1]/section[2]/div["+str(i)+"]/table/tbody/tr[6]/td")
            for wkincome in wkincomes:
                if len(wkincome.text) != 0:
                    incomes.append(wkincome.text)
                else:
                    incomes.append('')

        # Excel入力
        idx = 0
        for title in titles:
            ws.cell(rownumber, 1).value = title
            ws.cell(rownumber, 2).value = positiones[idx]
            ws.cell(rownumber, 3).value = wkplaces[idx]
            ws.cell(rownumber, 4).value = incomes[idx]
            idx += 1
            rownumber += 1

        wb.save(path)
        wb.close()

        # 次ページ移行
        nextbutton = driver.find_elements(By.CLASS_NAME, "link_next")
        if len(nextbutton)>1:
            nextbutton[0].click()
        else:
            break;

    # クローズ処理
    time.sleep(5)
    driver.close()
    driver.quit()

def ResourcePath(relativePath):
    try:
        basePath=sys._MEIPASS
    except Exception:
        basePath=os.path.dirname(__file__)
    return os.path.join(basePath, relativePath)

if __name__=="__main__":
    main()
